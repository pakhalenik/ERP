import openpyxl
from openpyxl.styles import Font

def create_excel():
    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Data for each order
    orders = {
        "Order 1": [
            ["Time (Minutes)", "Task", "Explanation"],
            ["0-6", "Wash & Mix", "Kristen prepares the dough for Order 1."],
            ["6-8", "Spoon Dough", "She spoons the dough onto the tray for Order 1."],
            ["8-9", "Load Oven", "She loads the tray into the oven for Order 1."],
            ["9-19", "Bake (automated)", "The cookies for Order 1 bake. Kristen can work on other tasks during this time."],
            ["19-24", "Cool (automated)", "The cookies cool. Kristen can work on other tasks during this time."],
            ["24-26", "Pack", "Kristen packs the cookies for Order 1."],
            ["26-27", "Accept Payment", "She accepts payment for Order 1."],
        ],
        "Order 2": [
            ["Time (Minutes)", "Task", "Explanation"],
            ["9-15", "Wash & Mix", "Kristen prepares the dough for Order 2 while Order 1 is baking."],
            ["15-17", "Spoon Dough", "She spoons the dough onto the tray for Order 2."],
            ["17-18", "Wait (oven busy)", "She waits for the oven to be free (Order 1 is still baking)."],
            ["18-19", "Load Oven", "She loads the tray into the oven for Order 2."],
            ["19-28", "Bake (automated)", "The cookies for Order 2 bake. Kristen can work on other tasks during this time."],
            ["28-33", "Cool (automated)", "The cookies cool. Kristen can work on other tasks during this time."],
            ["33-35", "Pack", "Kristen packs the cookies for Order 2."],
            ["35-36", "Accept Payment", "She accepts payment for Order 2."],
        ],
        "Order 3": [
            ["Time (Minutes)", "Task", "Explanation"],
            ["19-25", "Wash & Mix", "Kristen prepares the dough for Order 3 while Order 2 is baking."],
            ["25-27", "Spoon Dough", "She spoons the dough onto the tray for Order 3."],
            ["27-28", "Load Oven", "She loads the tray into the oven for Order 3."],
            ["28-38", "Bake (automated)", "The cookies for Order 3 bake. Kristen can work on other tasks during this time."],
            ["38-43", "Cool (automated)", "The cookies cool. Kristen can work on other tasks during this time."],
            ["38-40", "Pack", "Kristen packs the cookies for Order 3."],
            ["40-41", "Accept Payment", "She accepts payment for Order 3."],
        ],
        "Order 4": [
            ["Time (Minutes)", "Task", "Explanation"],
            ["28-34", "Wash & Mix", "Kristen prepares the dough for Order 4 while Order 3 is baking."],
            ["34-36", "Spoon Dough", "She spoons the dough onto the tray for Order 4."],
            ["36-37", "Wait (oven busy)", "She waits for the oven to be free (Order 3 is still baking)."],
            ["37-38", "Load Oven", "She loads the tray into the oven for Order 4."],
            ["38-47", "Bake (automated)", "The cookies for Order 4 bake. Kristen can work on other tasks during this time."],
            ["47-52", "Cool (automated)", "The cookies cool. Kristen can work on other tasks during this time."],
            ["47-49", "Pack", "Kristen packs the cookies for Order 4."],
            ["49-50", "Accept Payment", "She accepts payment for Order 4."],
        ],
    }

    # Create a sheet for each order
    for order, data in orders.items():
        sheet = workbook.create_sheet(title=order)
        for row in data:
            sheet.append(row)

    # Format headers for each sheet
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for col in worksheet[1]:
            col.font = Font(bold=True)

    # Create a summary sheet
    summary_data = [
        ["Order", "Start Time", "End Time", "Total Time"],
        ["Order 1", "0 minutes", "27 minutes", "27 minutes"],
        ["Order 2", "9 minutes", "36 minutes", "27 minutes"],
        ["Order 3", "19 minutes", "41 minutes", "22 minutes"],
        ["Order 4", "28 minutes", "50 minutes", "22 minutes"],
    ]

    summary_sheet = workbook.create_sheet(title="Summary")
    for row in summary_data:
        summary_sheet.append(row)

    # Format headers in summary sheet
    for col in summary_sheet[1]:
        col.font = Font(bold=True)

    # Save the workbook
    workbook.save("orders_schedule.xlsx")
    print("Excel file 'orders_schedule.xlsx' created successfully!")

# Run the function
create_excel()
